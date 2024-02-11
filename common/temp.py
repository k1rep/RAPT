import openpyxl

path = r'C:\Users\84926\Desktop\1104_2.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
wr = wb.active
row_list_total = [[i + 3, i + 18, i + 33, i + 48] for i in range(15)]
# row_list_total = [[i + 3, i + 18, i + 33] for i in range(15)]
# row_list_total = [[i + 3, i + 18] for i in range(15)]
col_list = [i + 3 for i in range(9)]
# total = [0 for i in range(2, 2 + 14 * 3)]
# result_col = 20
# font = openpyxl.styles.Font(name="微软雅黑",size=11,bold=True,color="FFA500")
fill = openpyxl.styles.PatternFill("solid", fgColor="FFA500")

for col in col_list:
    for row_list in row_list_total:
        sort_list = []
        for row in row_list:
            if wr.cell(row, col).value == None:
                break
            sort_list.append(wr.cell(row, col).value)
        if wr.cell(row, col).value == None:
            continue
        maxvalue = max(sort_list)
        for index in range(len(sort_list)):
            if sort_list[index] == maxvalue:
                # total[row_list[index] - 2] += 1
                wr.cell(row_list[index], col).fill = fill
    # for i in range(len(total)):
    #     wr.cell(i + 2, result_col).value = total[i]
wb.save(path)
